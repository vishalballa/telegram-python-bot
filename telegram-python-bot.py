import asyncio
import sys
import os
import datetime
import threading
import time
import csv
import json
import aiofiles
import uuid
import traceback
from decimal import Decimal
from collections import deque

import numpy as np
import pandas as pd

try:
    import pandas_ta as ta
    PANDAS_TA_AVAILABLE = True
except ImportError:
    ta = None
    PANDAS_TA_AVAILABLE = False
    print("WARNING: pandas_ta could not be imported. Technical indicators will be limited.")

from openpyxl import Workbook, load_workbook
from deriv_api import DerivAPI

try:
    from telegram import Bot
    from telegram.error import TelegramError
    TELEGRAM_AVAILABLE = True
except ImportError:
    Bot = None
    TelegramError = Exception
    TELEGRAM_AVAILABLE = False
    print("WARNING: python-telegram-bot could not be imported properly.")

from flask import Flask

# ==========================================
# 1. USER CONFIGURATION v10.2 (Divide & Rule Edition)
# ==========================================
CONFIG = {
    "APP_ID": int(os.environ.get("DERIV_APP_ID", "112872")),
    "API_TOKEN": os.environ.get("DERIV_API_TOKEN", "7QVkJpckczYk5ry"),
    "TELEGRAM_TOKEN": os.environ.get("TELEGRAM_TOKEN", "8382690806:AAElwVI3mujyE9T-zR-ZyT9ZJihb79Aj0mo"),
    "TELEGRAM_CHAT_ID": os.environ.get("TELEGRAM_CHAT_ID", "2145640324"),

    "VOLATILITY_ASSETS": ["R_10", "R_25", "R_50", "R_75", "R_100"],
    "TIMEFRAME_MAJOR": 300,
    "TIMEFRAME_ENTRY": 60,
    "SNIPER_DURATION": 15,
    "DURATION_UNIT": "m",

    "DAILY_TARGET_PROFIT": 50000.00,
    "DAILY_MAX_LOSS": 50000.00,

    "SNIPER_STAKE": 15.00,

    "ADX_THRESHOLD": 25,
    "ADX_CRASH_THRESHOLD": 40,

    "EXECUTION_DELAY": 1.0,
    "POST_TRADE_COOLDOWN": 300,
}

# ==========================================
# 1.1 DEFAULT CONFIG (For Stable Giants: R_50, R_75, R_100)
# Strategy: Let profits run with wide breathing room.
# ==========================================
DEFAULT_CONFIG = {
    "SAFETY_LOCK_PCT": 0.20,        # Lock at 20%
    "SAFETY_LOCK_SECURE_PCT": 0.05, # Secure 5%
    "SHADOW_ACTIVATION_PCT": 0.40,  # Start trailing at 40%
    "SHADOW_GAP_PCT": 0.35,         # WIDE GAP (35%)
    "TAKE_PROFIT_PCT": 0.80,        # HARD TARGET (80%) - Force Exit at $120 (on $150 stake)
    "SAFETY_STOP_PCT": 0.12,
    "GRACE_PERIOD_SEC": 30
}

# ==========================================
# 1.2 SYMBOL OVERRIDES (The "Divide & Rule" Logic)
# Strategy: Tight Leash for Thieves (R_10, R_25).
# Goal: If it spikes to $139, capture at least $125. Don't give it back!
# ==========================================
SYMBOL_CONFIG = {
    "R_10": {
        "SAFETY_LOCK_PCT": 0.15,        # Lock Early (15%)
        "SAFETY_LOCK_SECURE_PCT": 0.05, # Secure 5%
        "SHADOW_ACTIVATION_PCT": 0.25,  # Trail Early (25%)
        "SHADOW_GAP_PCT": 0.10,         # TIGHT GAP (10%)
        "TAKE_PROFIT_PCT": 0.65,        # LOWER TARGET (65%) - Take $97.5 and run
        "SAFETY_STOP_PCT": 0.12,
        "GRACE_PERIOD_SEC": 15
    },
    "R_25": {
        "SAFETY_LOCK_PCT": 0.15,        # Lock Early (15%)
        "SAFETY_LOCK_SECURE_PCT": 0.05, # Secure 5%
        "SHADOW_ACTIVATION_PCT": 0.25,  # Trail Early (25%)
        "SHADOW_GAP_PCT": 0.10,         # TIGHT GAP (10%)
        "TAKE_PROFIT_PCT": 0.65,        # LOWER TARGET (65%) - Take $97.5 and run
        "SAFETY_STOP_PCT": 0.12,
        "GRACE_PERIOD_SEC": 15
    }
}


def get_symbol_config(symbol):
    """Get configuration for a specific symbol (merges with defaults)"""
    config = DEFAULT_CONFIG.copy()
    if symbol in SYMBOL_CONFIG:
        config.update(SYMBOL_CONFIG[symbol])
    return config

FORENSICS_FILE = "trade_forensics.csv"
BLACKBOX_FILE = "blackbox_trades.csv"

def get_excel_file_for_symbol(symbol):
    """Get the Excel file path for a specific symbol (market-wise silos)"""
    return f"trades_{symbol}.xlsx"

# ==========================================
# 2. GLOBAL STATE v10.1 (Single Thread Discipline)
# ==========================================
class BotState:
    def __init__(self):
        self.total_pnl = 0.0
        self.wins = 0
        self.losses = 0
        self.trade_log = []
        self.daily_active = True
        self.early_sells = 0
        self.start_time = datetime.datetime.now()
        self.active_trades = {}
        self.symbol_locks = {}
        self.last_loss_time = {}
        self.cached_balance = 0.0
        self.cached_currency = "USD"
        self.balance_last_updated = 0.0
        self.paused_symbols = {}
        self.sniper_pnl = 0.0
        self.last_execution_time = 0.0
        self.safety_stop_count = 0
        self.timeout_count = 0
        self.pure_win_count = 0
        self.trend_surfer_wins = 0
        self.reversal_hunter_wins = 0
        self.crash_hunter_wins = 0
        self.shadow_exits = 0
        self.lock_exits = 0
        self.restart_count = 0
        self.last_heartbeat = time.time()
        self.last_trade_end_time = 0.0
        self.trade_in_progress = False

state = BotState()
balance_lock = asyncio.Lock()
execution_lock = asyncio.Lock()
trade_lock = asyncio.Lock()

trade_history_buffer = []
trade_buffer_lock = asyncio.Lock()

# ==========================================
# 2.1 CIRCULAR BUFFER - Price History (20 ticks)
# ==========================================
price_history_buffer = {}
price_buffer_lock = asyncio.Lock()

def init_price_history_buffer():
    for symbol in CONFIG["VOLATILITY_ASSETS"]:
        price_history_buffer[symbol] = deque(maxlen=20)
    print("[FORENSICS v10.0] Price history buffer initialized for all symbols (20 ticks each)")


# ==========================================
# 2.2 BLACK BOX FORENSICS v10.0 (Maximum Details)
# ==========================================
def init_blackbox_csv():
    if not os.path.exists(BLACKBOX_FILE):
        with open(BLACKBOX_FILE, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Trade_ID', 'Time', 'Symbol', 'Type', 'Strategy',
                'Entry_Price', 'Exit_Price', 'Profit_Loss', 'Duration_Secs',
                'RSI_Value', 'ADX_Value', 'EMA_Gap', 'Bollinger_Width',
                'Exit_Reason', 'Peak_Profit', 'Lock_Triggered', 'Shadow_Triggered',
                'Pre_Trade_Ticks', 'Post_Trade_Ticks'
            ])
            print(f"[BLACK BOX v10.0] Created {BLACKBOX_FILE}")


class BlackBoxForensics:
    def __init__(self, trade_id, symbol):
        self.trade_id = trade_id
        self.symbol = symbol
        self.pre_trade_ticks = []
        self.post_trade_ticks = []
        self.entry_technicals = {}
        self.csv_file = FORENSICS_FILE
        self._init_csv()

    def _init_csv(self):
        if not os.path.exists(self.csv_file):
            with open(self.csv_file, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Trade_ID', 'Symbol', 'Phase', 'Time', 'Price', 'Profit_PnL', 'Comment'])

    def capture_entry_technicals(self, df):
        if df is not None and len(df) > 0:
            current = df.iloc[-1]
            self.entry_technicals = {
                'rsi': float(current.get('rsi', 0)),
                'adx': float(current.get('adx', 0)),
                'ema50': float(current.get('ema50', 0)),
                'ema200': float(current.get('ema200', 0)),
                'close': float(current.get('close', 0)),
                'bb_upper': float(current.get('bb_upper', 0)),
                'bb_lower': float(current.get('bb_lower', 0)),
                'bb_middle': float(current.get('bb_middle', 0)),
            }
            self.entry_technicals['ema_gap'] = self.entry_technicals['close'] - self.entry_technicals['ema50']
            self.entry_technicals['bollinger_width'] = self.entry_technicals['bb_upper'] - self.entry_technicals['bb_lower']

    async def capture_pre_trade_ticks(self):
        async with price_buffer_lock:
            if self.symbol in price_history_buffer:
                self.pre_trade_ticks = list(price_history_buffer[self.symbol])

        if self.pre_trade_ticks:
            rows = []
            snapshot_len = len(self.pre_trade_ticks)
            for i, tick in enumerate(self.pre_trade_ticks):
                history_index = -(snapshot_len - i)
                rows.append([
                    self.trade_id,
                    self.symbol,
                    'PRE',
                    tick['time'],
                    tick['price'],
                    0.0,
                    f'History {history_index}'
                ])
            await self._write_rows_async(rows)
            print(f"[BLACK BOX v10.0] Captured {len(self.pre_trade_ticks)} PRE ticks for {self.symbol}")
        else:
            print(f"[BLACK BOX v10.0] WARNING: No buffer data for {self.symbol}")

    async def log_active_tick(self, price, profit, comment=""):
        row = [
            self.trade_id,
            self.symbol,
            'ACTIVE',
            datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3],
            price,
            profit,
            comment
        ]
        await self._write_row_async(row)

    async def capture_post_trade_ticks(self, tick_count=20):
        print(f"[BLACK BOX v10.0] Recording {tick_count} POST ticks for {self.symbol}...")

        for i in range(tick_count):
            await asyncio.sleep(1)
            try:
                async with price_buffer_lock:
                    if self.symbol in price_history_buffer and len(price_history_buffer[self.symbol]) > 0:
                        last_tick = price_history_buffer[self.symbol][-1]
                        self.post_trade_ticks.append(last_tick)
                        row = [
                            self.trade_id,
                            self.symbol,
                            'POST',
                            last_tick['time'],
                            last_tick['price'],
                            0.0,
                            f'After Exit +{i+1}s'
                        ]
                        await self._write_row_async(row)
            except Exception as e:
                print(f"[BLACK BOX v10.0] Post-trade tick error: {e}")

        print(f"[BLACK BOX v10.0] Captured {len(self.post_trade_ticks)} POST ticks for {self.symbol}")

    async def write_blackbox_record(self, trade_data):
        try:
            pre_ticks_json = json.dumps([t['price'] for t in self.pre_trade_ticks]) if self.pre_trade_ticks else "[]"
            post_ticks_json = json.dumps([t['price'] for t in self.post_trade_ticks]) if self.post_trade_ticks else "[]"

            row = [
                self.trade_id,
                trade_data.get('time', ''),
                self.symbol,
                trade_data.get('type', ''),
                trade_data.get('strategy', ''),
                trade_data.get('entry_price', 0),
                trade_data.get('exit_price', 0),
                trade_data.get('profit', 0),
                trade_data.get('duration_secs', 0),
                self.entry_technicals.get('rsi', 0),
                self.entry_technicals.get('adx', 0),
                self.entry_technicals.get('ema_gap', 0),
                self.entry_technicals.get('bollinger_width', 0),
                trade_data.get('exit_reason', ''),
                trade_data.get('peak_profit', 0),
                trade_data.get('lock_triggered', False),
                trade_data.get('shadow_triggered', False),
                pre_ticks_json,
                post_ticks_json
            ]

            async with aiofiles.open(BLACKBOX_FILE, 'a', newline='') as f:
                line = ','.join(str(x) for x in row) + '\n'
                await f.write(line)
            print(f"[BLACK BOX v10.0] Trade record saved: {self.trade_id}")
        except Exception as e:
            print(f"[BLACK BOX v10.0] Error writing blackbox record: {e}")

    async def _write_row_async(self, row):
        try:
            async with aiofiles.open(self.csv_file, 'a', newline='') as f:
                line = ','.join(str(x) for x in row) + '\n'
                await f.write(line)
        except Exception as e:
            print(f"[BLACK BOX v10.0] CSV write error: {e}")

    async def _write_rows_async(self, rows):
        try:
            async with aiofiles.open(self.csv_file, 'a', newline='') as f:
                lines = '\n'.join(','.join(str(x) for x in row) for row in rows) + '\n'
                await f.write(lines)
        except Exception as e:
            print(f"[BLACK BOX v10.0] CSV batch write error: {e}")


# ==========================================
# 3. STARTUP DIAGNOSTICS v10.0 (The Real Check)
# ==========================================
def print_dynamic_math_check():
    stake = CONFIG["SNIPER_STAKE"]
    print("=" * 60)
    print(f"🚀 AVITS BOT v11.1 | FULL AUTO-REPORTER EDITION")
    print(f"💰 STAKE: ${stake:.2f}")
    print("=" * 60)

    print("\n📊 PER-SYMBOL CONFIGURATION (v10.2 Divide & Rule):")
    print("-" * 60)

    print("\n🐢 STABLE GIANTS (R_50, R_75, R_100) - Trend Riding:")
    default_cfg = DEFAULT_CONFIG
    print(f"   - Lock Trigger  ({default_cfg['SAFETY_LOCK_PCT']*100:.0f}%): ${stake * default_cfg['SAFETY_LOCK_PCT']:.2f}")
    print(f"   - Secure Amount ({default_cfg['SAFETY_LOCK_SECURE_PCT']*100:.0f}%): ${stake * default_cfg['SAFETY_LOCK_SECURE_PCT']:.2f}")
    print(f"   - Shadow Trail  ({default_cfg['SHADOW_ACTIVATION_PCT']*100:.0f}%): ${stake * default_cfg['SHADOW_ACTIVATION_PCT']:.2f}")
    print(f"   - Shadow Gap    ({default_cfg['SHADOW_GAP_PCT']*100:.0f}%): ${stake * default_cfg['SHADOW_GAP_PCT']:.2f} [WIDE]")
    print(f"   - Safety Stop   ({default_cfg['SAFETY_STOP_PCT']*100:.0f}%): -${stake * default_cfg['SAFETY_STOP_PCT']:.2f}")
    print(f"   - Grace Period      : {default_cfg['GRACE_PERIOD_SEC']}s")

    print("\n⚡ VOLATILE THIEVES (R_10, R_25) - Aggressive Scalping:")
    volatile_cfg = SYMBOL_CONFIG.get("R_10", default_cfg)
    print(f"   - Lock Trigger  ({volatile_cfg['SAFETY_LOCK_PCT']*100:.0f}%): ${stake * volatile_cfg['SAFETY_LOCK_PCT']:.2f} [EARLY]")
    print(f"   - Secure Amount ({volatile_cfg['SAFETY_LOCK_SECURE_PCT']*100:.0f}%): ${stake * volatile_cfg['SAFETY_LOCK_SECURE_PCT']:.2f}")
    print(f"   - Shadow Trail  ({volatile_cfg['SHADOW_ACTIVATION_PCT']*100:.0f}%): ${stake * volatile_cfg['SHADOW_ACTIVATION_PCT']:.2f} [EARLY]")
    print(f"   - Shadow Gap    ({volatile_cfg['SHADOW_GAP_PCT']*100:.0f}%): ${stake * volatile_cfg['SHADOW_GAP_PCT']:.2f} [TIGHT]")
    print(f"   - Safety Stop   ({volatile_cfg['SAFETY_STOP_PCT']*100:.0f}%): -${stake * volatile_cfg['SAFETY_STOP_PCT']:.2f}")
    print(f"   - Grace Period      : {volatile_cfg['GRACE_PERIOD_SEC']}s [FAST]")

    print("\n" + "=" * 60)
    print("🔒 SINGLE THREAD DISCIPLINE (v10.1):")
    print(f"   - Max Concurrent Trades: 1")
    print(f"   - Post-Trade Cooldown: {CONFIG['POST_TRADE_COOLDOWN']}s (5 minutes)")
    print(f"   - One Bullet Policy: ACTIVE")
    print("=" * 60)


def run_diagnostics():
    print("\n" + "=" * 60)
    print("AVITS BOT v10.2 - DIVIDE & RULE EDITION")
    print("=" * 60)
    print("SYSTEM DIAGNOSTICS:")
    print("-" * 60)

    issues = []

    print(f"Mode:                 Pure Sniper (Single Thread)")
    print(f"")

    active_markets = CONFIG["VOLATILITY_ASSETS"]
    print(f"Markets (Cream Only): {', '.join(active_markets)}")
    print(f"Timeframes:           Major=5min (300s), Entry=1min (60s)")
    print(f"")

    print(f"[CRASH HUNTER SYSTEM]")
    print(f"  Normal Mode:        ADX < {CONFIG['ADX_CRASH_THRESHOLD']} (Strict BB Pullback)")
    print(f"  Crash Hunter:       ADX >= {CONFIG['ADX_CRASH_THRESHOLD']} (Override Safety)")
    print(f"")

    print(f"[ZOMBIE INFRASTRUCTURE]")
    print(f"  Heartbeat:          Every 60s")
    print(f"  Auto-Restart:       On crash (10s delay)")
    print(f"  Web Server:         Port 5000")
    print(f"")

    print(f"[BLACK BOX FORENSICS v10.0]")
    print(f"  PRE Phase:          20 ticks (captured before entry)")
    print(f"  ACTIVE Phase:       Every tick during trade")
    print(f"  POST Phase:         20 ticks after exit")
    print(f"  Data File:          {BLACKBOX_FILE}")
    print(f"")

    telegram_configured = bool(CONFIG["TELEGRAM_TOKEN"] and CONFIG["TELEGRAM_CHAT_ID"])
    telegram_status = "CONNECTED" if telegram_configured else "NOT CONFIGURED"
    print(f"Telegram Service:     {telegram_status}")
    if not telegram_configured:
        issues.append("Telegram not configured - alerts will only show in console")

    pandas_ta_status = "LOADED" if ta is not None else "FAILED"
    print(f"Pandas TA Library:    {pandas_ta_status}")
    if ta is None:
        issues.append("pandas_ta not available - technical analysis will be limited")

    deriv_token_status = "SET" if CONFIG["API_TOKEN"] else "MISSING"
    print(f"Deriv API Token:      {deriv_token_status}")
    if not CONFIG["API_TOKEN"]:
        issues.append("DERIV_API_TOKEN is missing - bot cannot trade")

    deriv_app_id_status = "SET" if CONFIG["APP_ID"] else "MISSING"
    print(f"Deriv App ID:         {deriv_app_id_status}")

    print("-" * 60)

    if issues:
        print("WARNINGS:")
        for issue in issues:
            print(f"  - {issue}")
        print("-" * 60)

    if CONFIG["API_TOKEN"] and ta is not None:
        print("DIAGNOSTICS PASSED. STARTING BOT...")
    elif CONFIG["API_TOKEN"]:
        print("DIAGNOSTICS PASSED WITH WARNINGS. STARTING BOT...")
    else:
        print("DIAGNOSTICS FAILED. MISSING CRITICAL SECRETS.")
        print("\nPlease set the following environment variables:")
        print("  - DERIV_API_TOKEN: Your Deriv API token")
        print("  - DERIV_APP_ID: Your Deriv App ID")

    print("=" * 60 + "\n")

    return len(issues) == 0 or CONFIG["API_TOKEN"]


# ==========================================
# 4. EXCEL LOGGING SYSTEM (Enhanced v11.1 - Market-Wise Silos)
# ==========================================
def init_excel_for_symbol(symbol):
    """Initialize Excel file for a specific symbol (market-wise silos)"""
    excel_file = get_excel_file_for_symbol(symbol)
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Trades"
            ws.append([
                'Time', 'Trade_ID', 'Symbol', 'Strategy', 'Type', 'Stake',
                'Entry_Price', 'Exit_Price', 'Profit', 'Result',
                'Duration_Secs', 'Exit_Reason', 'Peak_Profit',
                'RSI', 'ADX', 'EMA_Gap', 'BB_Width',
                'Lock_Triggered', 'Shadow_Triggered',
                'Slippage', 'Latency_ms', 'Exit_RSI', 'Exit_ADX'
            ])
            wb.save(excel_file)
            print(f"[MARKET SILO v11.1] Created {excel_file}")


def init_excel():
    """Initialize Excel files for all configured symbols"""
    for symbol in CONFIG["VOLATILITY_ASSETS"]:
        init_excel_for_symbol(symbol)


def log_trade_to_excel(trade_data):
    """Log trade to symbol-specific Excel file with forensic data"""
    try:
        symbol = trade_data.get('symbol', 'UNKNOWN')
        excel_file = get_excel_file_for_symbol(symbol)
        
        if not os.path.exists(excel_file):
            init_excel_for_symbol(symbol)
        
        wb = load_workbook(excel_file)
        ws = wb.active
        if ws is not None:
            ws.append([
                trade_data.get('time', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                trade_data.get('trade_id', ''),
                trade_data.get('symbol', ''),
                trade_data.get('strategy', ''),
                trade_data.get('type', ''),
                trade_data.get('stake', 0),
                trade_data.get('entry_price', 0),
                trade_data.get('exit_price', 0),
                trade_data.get('profit', 0),
                trade_data.get('result', ''),
                trade_data.get('duration_secs', 0),
                trade_data.get('exit_reason', ''),
                trade_data.get('peak_profit', 0),
                trade_data.get('rsi', 0),
                trade_data.get('adx', 0),
                trade_data.get('ema_gap', 0),
                trade_data.get('bb_width', 0),
                trade_data.get('lock_triggered', False),
                trade_data.get('shadow_triggered', False),
                trade_data.get('slippage', 0),
                trade_data.get('latency_ms', 0),
                trade_data.get('exit_rsi', 0),
                trade_data.get('exit_adx', 0)
            ])
            wb.save(excel_file)
            print(f"[MARKET SILO v11.1] Trade logged to {excel_file}")
    except Exception as e:
        print(f"Excel logging error: {e}")


# ==========================================
# 5. FLASK KEEP-ALIVE SERVER
# ==========================================
app = Flask(__name__)


@app.route('/')
def home():
    uptime = datetime.datetime.now() - state.start_time
    hours, remainder = divmod(int(uptime.total_seconds()), 3600)
    minutes, seconds = divmod(remainder, 60)
    active_count = len(state.active_trades)
    total_trades = state.wins + state.losses
    win_rate = (state.wins / total_trades * 100) if total_trades > 0 else 0
    stake = CONFIG["SNIPER_STAKE"]

    cooldown_remaining = max(0, CONFIG["POST_TRADE_COOLDOWN"] - (time.time() - state.last_trade_end_time))
    cooldown_status = f"☕ {int(cooldown_remaining)}s" if cooldown_remaining > 0 else "✅ Ready"

    volatile_cfg = SYMBOL_CONFIG.get("R_10", DEFAULT_CONFIG)
    stable_cfg = DEFAULT_CONFIG

    return f"""
    <html>
    <head>
        <title>AVITS Trading Bot v10.2</title>
        <meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate" />
        <meta http-equiv="refresh" content="10">
        <style>
            body {{ font-family: Arial, sans-serif; padding: 20px; background: #1a1a2e; color: #eee; }}
            h1 {{ color: #00ff88; }}
            .stat {{ margin: 10px 0; padding: 10px; background: #16213e; border-radius: 5px; }}
            .profit {{ color: #00ff88; }}
            .loss {{ color: #ff4444; }}
            .zombie {{ color: #ff00ff; }}
            .dynamic {{ color: #ffaa00; }}
            .crash {{ color: #00ffff; }}
            .discipline {{ color: #ff6600; }}
            .volatile {{ color: #ff4488; }}
            .stable {{ color: #44ff88; }}
            .config-table {{ width: 100%; border-collapse: collapse; margin: 10px 0; }}
            .config-table th, .config-table td {{ border: 1px solid #333; padding: 8px; text-align: left; }}
            .config-table th {{ background: #0f3460; }}
        </style>
    </head>
    <body>
        <h1>AVITS Trading Bot v10.2 - Divide & Rule Edition</h1>
        <div class="stat zombie">ZOMBIE MODE: ACTIVE | Restarts: {state.restart_count}</div>
        <div class="stat discipline">🔒 SINGLE THREAD: {'⏳ Trade Active' if active_count > 0 else cooldown_status}</div>
        <div class="stat">Status: {'🟢 Active' if state.daily_active else '🔴 Stopped'}</div>
        <div class="stat">Total P/L: <span class="{'profit' if state.total_pnl >= 0 else 'loss'}">${state.total_pnl:.2f}</span></div>
        <div class="stat">Sniper P/L: ${state.sniper_pnl:.2f}</div>
        <div class="stat dynamic">Dynamic Stake: ${stake:.2f}</div>

        <h3>Per-Symbol Configuration (Divide & Rule)</h3>
        <table class="config-table">
            <tr>
                <th>Setting</th>
                <th class="volatile">R_10, R_25 (Volatile)</th>
                <th class="stable">R_50, R_75, R_100 (Stable)</th>
            </tr>
            <tr>
                <td>Lock Trigger</td>
                <td class="volatile">{volatile_cfg['SAFETY_LOCK_PCT']*100:.0f}% (${stake*volatile_cfg['SAFETY_LOCK_PCT']:.2f})</td>
                <td class="stable">{stable_cfg['SAFETY_LOCK_PCT']*100:.0f}% (${stake*stable_cfg['SAFETY_LOCK_PCT']:.2f})</td>
            </tr>
            <tr>
                <td>Shadow Trail</td>
                <td class="volatile">{volatile_cfg['SHADOW_ACTIVATION_PCT']*100:.0f}% (${stake*volatile_cfg['SHADOW_ACTIVATION_PCT']:.2f})</td>
                <td class="stable">{stable_cfg['SHADOW_ACTIVATION_PCT']*100:.0f}% (${stake*stable_cfg['SHADOW_ACTIVATION_PCT']:.2f})</td>
            </tr>
            <tr>
                <td>Shadow Gap</td>
                <td class="volatile">{volatile_cfg['SHADOW_GAP_PCT']*100:.0f}% (${stake*volatile_cfg['SHADOW_GAP_PCT']:.2f}) [TIGHT]</td>
                <td class="stable">{stable_cfg['SHADOW_GAP_PCT']*100:.0f}% (${stake*stable_cfg['SHADOW_GAP_PCT']:.2f}) [WIDE]</td>
            </tr>
            <tr>
                <td>Grace Period</td>
                <td class="volatile">{volatile_cfg['GRACE_PERIOD_SEC']}s [FAST]</td>
                <td class="stable">{stable_cfg['GRACE_PERIOD_SEC']}s</td>
            </tr>
        </table>

        <div class="stat crash">Crash Hunter: ADX >= {CONFIG['ADX_CRASH_THRESHOLD']} | Wins: {state.crash_hunter_wins}</div>
        <div class="stat">Wins: {state.wins} | Losses: {state.losses} | Win Rate: {win_rate:.1f}%</div>
        <div class="stat">Trend Surfer: {state.trend_surfer_wins} | Reversal: {state.reversal_hunter_wins}</div>
        <div class="stat">Exit Stats: Shadow: {state.shadow_exits} | Lock: {state.lock_exits} | Safety Stop: {state.safety_stop_count}</div>
        <div class="stat">Active Trades: {active_count}</div>
        <div class="stat">Uptime: {hours}h {minutes}m {seconds}s</div>
        <div class="stat">Markets: {', '.join(CONFIG['VOLATILITY_ASSETS'])}</div>
        <div class="stat">Last updated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
    </body>
    </html>
    """


@app.route('/health')
def health():
    return "OK", 200


@app.route('/heartbeat')
def heartbeat():
    state.last_heartbeat = time.time()
    return f"ALIVE | Uptime: {datetime.datetime.now() - state.start_time}", 200


def run_keep_alive():
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)


# ==========================================
# 5.1 HEARTBEAT THREAD (Zombie Infrastructure)
# ==========================================
def start_heartbeat():
    while True:
        try:
            uptime = datetime.datetime.now() - state.start_time
            hours, remainder = divmod(int(uptime.total_seconds()), 3600)
            minutes, _ = divmod(remainder, 60)
            active = len(state.active_trades)
            cooldown = max(0, CONFIG["POST_TRADE_COOLDOWN"] - (time.time() - state.last_trade_end_time))
            status = f"TRADING" if active > 0 else (f"COOLDOWN {int(cooldown)}s" if cooldown > 0 else "SCANNING")
            print(f"💓 BOT ALIVE | {status} | Uptime: {hours}h {minutes}m | Trades: {state.wins + state.losses} | P/L: ${state.total_pnl:.2f}")
            state.last_heartbeat = time.time()
            time.sleep(60)
        except Exception as e:
            print(f"Heartbeat error: {e}")
            time.sleep(60)


# ==========================================
# 6. PANDAS_TA INDICATOR ENGINE
# ==========================================
def calculate_indicators(df):
    if df is None or len(df) < 30:
        return None

    if ta is None:
        df['macd'] = 0
        df['macd_signal'] = 0
        df['macd_hist'] = 0
        df['rsi'] = 50
        df['atr'] = 0
        df['atr_sma'] = 0
        df['ema50'] = df['close']
        df['ema200'] = df['close']
        df['bb_lower'] = df['close']
        df['bb_middle'] = df['close']
        df['bb_upper'] = df['close']
        df['adx'] = 0
        return df

    try:
        close = df['close'].astype(float)
        high = df['high'].astype(float)
        low = df['low'].astype(float)

        macd_result = ta.macd(close, fast=12, slow=26, signal=9)
        if macd_result is not None and len(macd_result.columns) >= 3:
            df['macd'] = macd_result.iloc[:, 0]
            df['macd_hist'] = macd_result.iloc[:, 1]
            df['macd_signal'] = macd_result.iloc[:, 2]
        else:
            df['macd'] = 0
            df['macd_signal'] = 0
            df['macd_hist'] = 0

        rsi_result = ta.rsi(close, length=14)
        df['rsi'] = rsi_result if rsi_result is not None else 50

        atr_result = ta.atr(high, low, close, length=14)
        df['atr'] = atr_result if atr_result is not None else 0

        atr_sma = ta.sma(df['atr'], length=20)
        df['atr_sma'] = atr_sma if atr_sma is not None else 0

        ema50_result = ta.ema(close, length=50)
        df['ema50'] = ema50_result if ema50_result is not None else close

        ema200_result = ta.ema(close, length=200)
        df['ema200'] = ema200_result if ema200_result is not None else close

        bb_result = ta.bbands(close, length=20, std=2)
        if bb_result is not None and len(bb_result.columns) >= 3:
            df['bb_lower'] = bb_result.iloc[:, 0]
            df['bb_middle'] = bb_result.iloc[:, 1]
            df['bb_upper'] = bb_result.iloc[:, 2]
        else:
            df['bb_lower'] = close
            df['bb_middle'] = close
            df['bb_upper'] = close

        adx_result = ta.adx(high, low, close, length=14)
        if adx_result is not None and len(adx_result.columns) >= 1:
            df['adx'] = adx_result.iloc[:, 0]
        else:
            df['adx'] = 0

        return df

    except Exception as e:
        print(f"Indicator calculation error: {e}")
        return None


# ==========================================
# 7. TRADE MANAGER CLASS v10.0 (Black Box Edition)
# ==========================================
class TradeManager:

    def __init__(self, api, symbol, strategy_type, direction, stake, duration,
                 current_price=0.0, signal_strength=50, signal_name="UNKNOWN", 
                 is_crash_hunter=False, entry_df=None, signal_time=None, signal_price=None):
        self.api = api
        self.symbol = symbol
        self.strategy_type = strategy_type
        self.direction = direction
        self.stake = stake
        self.duration = duration
        self.current_price = current_price
        self.signal_strength = signal_strength
        self.signal_name = signal_name
        self.is_crash_hunter = is_crash_hunter
        self.entry_df = entry_df
        self.signal_time = signal_time if signal_time else time.time()
        self.signal_price = signal_price if signal_price else current_price
        self.contract_id = None
        self.buy_price = 0.0
        self.is_active = False
        self.result = None
        self.start_time = 0.0
        self.highest_profit = 0.0
        self.is_selling = False
        self.sell_reason_logged = False
        self.shadow_active = False
        self.lock_active = False
        self.trade_id = f"{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{str(uuid.uuid4())[:4]}"
        self.forensics = None
        self.actual_entry_price = 0.0
        self.exit_technicals = {'rsi': 0, 'adx': 0}

        sym_config = get_symbol_config(symbol)
        self.val_lock_trigger = self.stake * sym_config["SAFETY_LOCK_PCT"]
        self.val_lock_secure = self.stake * sym_config["SAFETY_LOCK_SECURE_PCT"]
        self.val_shadow_active = self.stake * sym_config["SHADOW_ACTIVATION_PCT"]
        self.val_shadow_gap = self.stake * sym_config["SHADOW_GAP_PCT"]
        self.val_take_profit = self.stake * sym_config["TAKE_PROFIT_PCT"]
        self.val_safety_stop = self.stake * sym_config["SAFETY_STOP_PCT"]
        self.grace_period = sym_config["GRACE_PERIOD_SEC"]
        self.sym_config = sym_config

    async def execute(self, currency=None):
        try:
            async with execution_lock:
                elapsed = time.time() - state.last_execution_time
                if elapsed < CONFIG["EXECUTION_DELAY"]:
                    wait_time = CONFIG["EXECUTION_DELAY"] - elapsed
                    print(f"Execution queue: waiting {wait_time:.2f}s...")
                    await asyncio.sleep(wait_time)
                state.last_execution_time = time.time()

            if self.symbol in state.active_trades:
                print(f"SKIP: {self.symbol} already has an active trade")
                return {"success": False, "profit": 0.0, "error": "Symbol already has active trade"}

            state.active_trades[self.symbol] = "PENDING"
            state.trade_in_progress = True

            if currency is None:
                _, currency = await get_cached_balance(self.api)

            mode = "CRASH HUNTER" if self.is_crash_hunter else "NORMAL"
            config_type = "TIGHT" if self.symbol in SYMBOL_CONFIG else "WIDE"
            print(f"\n{'='*60}")
            print(f"🎯 SNIPER v10.2 [{mode}] | Symbol: {self.symbol} | Config: {config_type}")
            print(f"📊 Dynamic: Lock@${self.val_lock_trigger:.2f} | Shadow@${self.val_shadow_active:.2f} | Gap@${self.val_shadow_gap:.2f} | Stop@-${self.val_safety_stop:.2f}")
            print(f"{'='*60}")

            self.forensics = BlackBoxForensics(self.trade_id, self.symbol)

            if self.entry_df is not None:
                self.forensics.capture_entry_technicals(self.entry_df)

            print(f"[BLACK BOX v10.0] Recording trade data... Trade ID: {self.trade_id}")
            await self.forensics.capture_pre_trade_ticks()

            result = await self._execute_sniper(currency)

            if not result.get('success', False):
                if self.symbol in state.active_trades and state.active_trades[self.symbol] == "PENDING":
                    del state.active_trades[self.symbol]
                state.trade_in_progress = False

            return result

        except Exception as e:
            if self.symbol in state.active_trades and state.active_trades[self.symbol] == "PENDING":
                del state.active_trades[self.symbol]
            state.trade_in_progress = False
            print(f"Trade execution error for {self.symbol}: {e}")
            return {"success": False, "profit": 0.0, "error": str(e)}

    async def _check_contract_available(self, contract_type):
        try:
            contracts = await self.api.contracts_for({"contracts_for": self.symbol})
            if 'contracts_for' in contracts:
                available = contracts['contracts_for'].get('available', [])
                for contract in available:
                    if contract.get('contract_type') == contract_type:
                        return True
        except Exception as e:
            if 'MarketIsClosed' in str(e) or 'SymbolNotAvailable' in str(e):
                print(f"Market closed or symbol not available for {self.symbol}")
            else:
                print(f"Contract check error for {self.symbol}: {e}")
        return False

    async def _execute_sniper(self, currency):
        try:
            contract_type = "CALL" if self.direction == "CALL" else "PUT"

            if not await self._check_contract_available(contract_type):
                print(f"Contract {contract_type} not available for {self.symbol}")
                return {"success": False, "profit": 0.0, "error": "Contract not available"}

            proposal_params = {
                "proposal": 1,
                "amount": self.stake,
                "basis": "stake",
                "contract_type": contract_type,
                "currency": currency,
                "duration": CONFIG["SNIPER_DURATION"],
                "duration_unit": CONFIG["DURATION_UNIT"],
                "symbol": self.symbol
            }

            print(f"[SNIPER v10.0] Requesting {contract_type} proposal for {self.symbol}...")
            proposal_response = await self.api.proposal(proposal_params)

            if 'error' in proposal_response:
                error_msg = proposal_response['error'].get('message', 'Unknown error')
                print(f"Proposal error: {error_msg}")
                return {"success": False, "profit": 0.0, "error": error_msg}

            proposal_id = proposal_response.get('proposal', {}).get('id')
            if not proposal_id:
                print("No proposal ID received")
                return {"success": False, "profit": 0.0, "error": "No proposal ID"}

            buy_params = {"buy": proposal_id, "price": self.stake}
            buy_response = await self.api.buy(buy_params)

            if 'error' in buy_response:
                error_msg = buy_response['error'].get('message', 'Unknown error')
                print(f"Buy error: {error_msg}")
                return {"success": False, "profit": 0.0, "error": error_msg}

            buy_info = buy_response.get('buy', {})
            self.contract_id = buy_info.get('contract_id')
            self.buy_price = float(buy_info.get('buy_price', 0))

            if not self.contract_id:
                print("No contract ID received")
                return {"success": False, "profit": 0.0, "error": "No contract ID"}

            self.is_active = True
            self.start_time = time.time()
            self.actual_entry_price = float(buy_info.get('entry_spot', buy_info.get('start_spot', self.current_price)))
            state.active_trades[self.symbol] = self.contract_id

            mode_tag = "🔥CRASH" if self.is_crash_hunter else "⚡SNIPER"
            print(f"[{mode_tag}] ENTRY: {self.symbol} {contract_type} | Contract: {self.contract_id} | Buy: ${self.buy_price:.2f}")

            await send_telegram(
                f"{mode_tag} ENTRY\n"
                f"Symbol: {self.symbol}\n"
                f"Direction: {contract_type}\n"
                f"Stake: ${self.stake:.2f}\n"
                f"Strategy: {self.strategy_type}\n"
                f"Signal: {self.signal_name}"
            )

            result = await self._monitor_sniper()
            return result

        except Exception as e:
            print(f"Sniper execution error: {e}")
            traceback.print_exc()
            return {"success": False, "profit": 0.0, "error": str(e)}

    async def _monitor_sniper(self):
        try:
            max_duration = CONFIG["SNIPER_DURATION"] * 60 + 120
            start_time = time.time()
            last_profit = 0.0
            exit_reason = "EXPIRED"

            while self.is_active and (time.time() - start_time) < max_duration:
                try:
                    poc_response = await self.api.proposal_open_contract({
                        "proposal_open_contract": 1,
                        "contract_id": self.contract_id
                    })

                    if 'error' in poc_response:
                        print(f"POC error: {poc_response['error']}")
                        await asyncio.sleep(1)
                        continue

                    poc = poc_response.get('proposal_open_contract', {})

                    if not poc:
                        await asyncio.sleep(0.5)
                        continue

                    current_profit = float(poc.get('profit', 0))
                    current_price = float(poc.get('current_spot', 0))
                    is_sold = poc.get('is_sold', 0) == 1
                    is_expired = poc.get('is_expired', 0) == 1

                    if self.forensics:
                        await self.forensics.log_active_tick(current_price, current_profit, f"Peak: {self.highest_profit:.2f}")

                    if current_profit > self.highest_profit:
                        self.highest_profit = current_profit

                    if is_sold or is_expired:
                        final_profit = float(poc.get('profit', 0))
                        sell_price = float(poc.get('sell_price', 0))
                        exit_reason = "EXPIRED" if is_expired else "CONTRACT_SOLD"

                        return await self._finalize_trade(final_profit, sell_price, exit_reason)

                    trade_duration = time.time() - self.start_time

                    should_exit, exit_reason = self._check_dynamic_exit(current_profit, trade_duration)

                    if should_exit and not self.is_selling:
                        self.is_selling = True
                        print(f"[v10.0 EXIT] {exit_reason} | Profit: ${current_profit:.2f}")

                        try:
                            sell_response = await self.api.sell({
                                "sell": self.contract_id,
                                "price": 0
                            })

                            if 'error' not in sell_response:
                                sell_info = sell_response.get('sell', {})
                                final_profit = float(sell_info.get('profit', current_profit))
                                sell_price = float(sell_info.get('sold_for', 0))
                                return await self._finalize_trade(final_profit, sell_price, exit_reason)
                        except Exception as sell_error:
                            print(f"Sell error: {sell_error}")
                            self.is_selling = False

                    last_profit = current_profit
                    await asyncio.sleep(0.5)

                except Exception as e:
                    print(f"Monitor error: {e}")
                    await asyncio.sleep(1)

            return await self._finalize_trade(last_profit, 0, "TIMEOUT")

        except Exception as e:
            print(f"Monitor fatal error: {e}")
            traceback.print_exc()
            return {"success": False, "profit": 0.0, "error": str(e)}

    def _check_dynamic_exit(self, profit, duration):
        """v11.1 Hybrid Sniper Exit - STRICT PRIORITY ORDER"""
        
        # 1. HARD TAKE PROFIT (TOP PRIORITY) - Prevent Flash Crashes
        if profit >= self.val_take_profit:
            state.pure_win_count += 1
            print(f"[HARD TAKE PROFIT] Profit ${profit:.2f} >= ${self.val_take_profit:.2f} - IMMEDIATE EXIT!")
            return True, "HARD_TAKE_PROFIT"
        
        # 2. SAFETY STOP - Check for Grace Period or Hard Stop
        if duration < self.grace_period:
            if profit <= -self.val_safety_stop:
                return True, "SAFETY_STOP_GRACE"
            return False, ""

        if profit <= -self.val_safety_stop:
            return True, "HARD_STOP"

        # 3. TRAILING STOP (SHADOW MODE)
        if profit >= self.val_shadow_active:
            if not self.shadow_active:
                self.shadow_active = True
                print(f"[SHADOW ACTIVATED] Profit ${profit:.2f} >= ${self.val_shadow_active:.2f}")

        if self.shadow_active:
            trailing_floor = self.highest_profit - self.val_shadow_gap
            if profit <= trailing_floor:
                state.shadow_exits += 1
                return True, "SHADOW_TRAIL_EXIT"

        # 4. SAFETY LOCK
        if profit >= self.val_lock_trigger:
            if not self.lock_active:
                self.lock_active = True
                print(f"[LOCK ACTIVATED] Profit ${profit:.2f} >= ${self.val_lock_trigger:.2f}")

        if self.lock_active:
            if profit <= self.val_lock_secure:
                state.lock_exits += 1
                return True, "SAFETY_LOCK_HIT"

        return False, ""

    async def _fetch_exit_technicals(self):
        """Fetch latest candle data after exit for forensic analysis"""
        try:
            candles = await self.api.ticks_history({
                "ticks_history": self.symbol,
                "adjust_start_time": 1,
                "count": 50,
                "end": "latest",
                "granularity": CONFIG["TIMEFRAME_ENTRY"],
                "style": "candles"
            })
            
            if 'candles' in candles and len(candles['candles']) > 0:
                df = pd.DataFrame(candles['candles'])
                df = calculate_indicators(df)
                if df is not None and len(df) > 0:
                    current = df.iloc[-1]
                    self.exit_technicals = {
                        'rsi': float(current.get('rsi', 0)),
                        'adx': float(current.get('adx', 0))
                    }
                    print(f"[FORENSICS v11.1] Exit Technicals: RSI={self.exit_technicals['rsi']:.1f}, ADX={self.exit_technicals['adx']:.1f}")
        except Exception as e:
            print(f"[FORENSICS v11.1] Exit technicals fetch error: {e}")

    async def _finalize_trade(self, profit, sell_price, exit_reason):
        try:
            self.is_active = False
            self.result = profit

            if self.symbol in state.active_trades:
                del state.active_trades[self.symbol]

            state.total_pnl += profit
            state.sniper_pnl += profit
            state.last_trade_end_time = time.time()
            state.trade_in_progress = False

            if profit >= 0:
                state.wins += 1
                if self.is_crash_hunter:
                    state.crash_hunter_wins += 1
                elif "TREND" in self.strategy_type.upper():
                    state.trend_surfer_wins += 1
                else:
                    state.reversal_hunter_wins += 1
                result_str = "WIN"
            else:
                state.losses += 1
                result_str = "LOSS"
                if "STOP" in exit_reason or "HARD" in exit_reason:
                    state.safety_stop_count += 1

            duration_secs = time.time() - self.start_time
            
            await self._fetch_exit_technicals()
            
            slippage = self.actual_entry_price - self.signal_price if self.actual_entry_price and self.signal_price else 0
            latency_ms = (self.start_time - self.signal_time) * 1000 if self.signal_time else 0

            trade_data = {
                'time': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'trade_id': self.trade_id,
                'symbol': self.symbol,
                'strategy': self.strategy_type,
                'type': self.direction,
                'stake': self.stake,
                'entry_price': self.current_price,
                'exit_price': sell_price,
                'profit': profit,
                'result': result_str,
                'duration_secs': int(duration_secs),
                'exit_reason': exit_reason,
                'peak_profit': self.highest_profit,
                'rsi': self.forensics.entry_technicals.get('rsi', 0) if self.forensics else 0,
                'adx': self.forensics.entry_technicals.get('adx', 0) if self.forensics else 0,
                'ema_gap': self.forensics.entry_technicals.get('ema_gap', 0) if self.forensics else 0,
                'bb_width': self.forensics.entry_technicals.get('bollinger_width', 0) if self.forensics else 0,
                'lock_triggered': self.lock_active,
                'shadow_triggered': self.shadow_active,
                'slippage': slippage,
                'latency_ms': latency_ms,
                'exit_rsi': self.exit_technicals.get('rsi', 0),
                'exit_adx': self.exit_technicals.get('adx', 0)
            }
            log_trade_to_excel(trade_data)

            emoji = "✅" if profit >= 0 else "❌"
            mode_tag = "CRASH" if self.is_crash_hunter else "SNIPER"
            print(f"\n[{mode_tag} {emoji}] {self.symbol} | {exit_reason} | Profit: ${profit:.2f} | Peak: ${self.highest_profit:.2f}")
            print(f"[FORENSICS v11.1] Slippage: {slippage:.4f} | Latency: {latency_ms:.1f}ms")
            print(f"Session: ${state.total_pnl:.2f} | W:{state.wins} L:{state.losses}")
            print(f"☕ Starting {CONFIG['POST_TRADE_COOLDOWN']}s cooldown...")

            await send_telegram(
                f"{emoji} {mode_tag} EXIT\n"
                f"Symbol: {self.symbol}\n"
                f"Result: {result_str}\n"
                f"Profit: ${profit:.2f}\n"
                f"Peak: ${self.highest_profit:.2f}\n"
                f"Exit: {exit_reason}\n"
                f"Duration: {int(duration_secs)}s\n"
                f"Latency: {latency_ms:.1f}ms\n"
                f"Session P/L: ${state.total_pnl:.2f}\n"
                f"Cooldown: {CONFIG['POST_TRADE_COOLDOWN']}s"
            )

            if self.forensics:
                await self.forensics.capture_post_trade_ticks(20)

                blackbox_data = trade_data.copy()
                blackbox_data['lock_triggered'] = self.lock_active
                blackbox_data['shadow_triggered'] = self.shadow_active
                await self.forensics.write_blackbox_record(blackbox_data)

            return {"success": True, "profit": profit, "exit_reason": exit_reason}

        except Exception as e:
            print(f"Finalize error: {e}")
            state.trade_in_progress = False
            return {"success": False, "profit": profit, "error": str(e)}


# ==========================================
# 8. SIGNAL DETECTION v10.0 (With Technical Data)
# ==========================================
def determine_sniper_signal(df, symbol):
    if df is None or len(df) < 30:
        return None, 0, "NO_DATA", None

    try:
        current = df.iloc[-1]
        prev = df.iloc[-2]
        prev2 = df.iloc[-3] if len(df) >= 3 else prev

        close = float(current['close'])
        ema50 = float(current['ema50'])
        bb_lower = float(current['bb_lower'])
        bb_upper = float(current['bb_upper'])
        rsi = float(current['rsi'])
        adx = float(current['adx'])

        prev_close = float(prev['close'])
        prev_open = float(prev.get('open', prev_close))
        prev2_close = float(prev2['close'])

        is_green_candle = prev_close > prev_open
        is_red_candle = prev_close < prev_open
        making_new_lows = close < prev_close < prev2_close
        making_new_highs = close > prev_close > prev2_close

        is_crash_mode = adx >= CONFIG["ADX_CRASH_THRESHOLD"]

        if is_crash_mode:
            if close < ema50 and making_new_lows:
                return "PUT", 85, f"CRASH_HUNTER_PUT (ADX:{adx:.1f})", df

            if close > ema50 and making_new_highs:
                return "CALL", 85, f"CRASH_HUNTER_CALL (ADX:{adx:.1f})", df

        if adx < CONFIG["ADX_THRESHOLD"]:
            return None, 0, f"ADX_TOO_LOW ({adx:.1f})", None

        if close <= bb_lower * 1.002 and rsi < 35 and is_green_candle:
            return "CALL", 75, f"NORMAL_PULLBACK_CALL (RSI:{rsi:.1f}, ADX:{adx:.1f})", df

        if close >= bb_upper * 0.998 and rsi > 65 and is_red_candle:
            return "PUT", 75, f"NORMAL_PULLBACK_PUT (RSI:{rsi:.1f}, ADX:{adx:.1f})", df

        if close > ema50 and prev_close <= ema50 and is_green_candle and rsi > 50:
            return "CALL", 70, f"TREND_SURFER_CALL (RSI:{rsi:.1f})", df

        if close < ema50 and prev_close >= ema50 and is_red_candle and rsi < 50:
            return "PUT", 70, f"TREND_SURFER_PUT (RSI:{rsi:.1f})", df

        return None, 0, "NO_SIGNAL", None

    except Exception as e:
        print(f"Signal detection error for {symbol}: {e}")
        return None, 0, "ERROR", None


# ==========================================
# 9. TELEGRAM NOTIFICATIONS
# ==========================================
async def send_telegram(message):
    if not TELEGRAM_AVAILABLE or not CONFIG["TELEGRAM_TOKEN"] or not CONFIG["TELEGRAM_CHAT_ID"]:
        return

    try:
        bot = Bot(token=CONFIG["TELEGRAM_TOKEN"])
        await bot.send_message(chat_id=CONFIG["TELEGRAM_CHAT_ID"], text=message)
    except Exception as e:
        print(f"Telegram error: {e}")


async def send_telegram_document(file_path, caption=""):
    """Send a document file via Telegram"""
    if not TELEGRAM_AVAILABLE or not CONFIG["TELEGRAM_TOKEN"] or not CONFIG["TELEGRAM_CHAT_ID"]:
        return False

    try:
        bot = Bot(token=CONFIG["TELEGRAM_TOKEN"])
        with open(file_path, 'rb') as doc:
            await bot.send_document(
                chat_id=CONFIG["TELEGRAM_CHAT_ID"],
                document=doc,
                caption=caption
            )
        return True
    except Exception as e:
        print(f"Telegram document error: {e}")
        return False


# ==========================================
# 9.1 AUTO-REPORTER DAEMON (v11.1 - Cloud Backups)
# ==========================================
def start_periodic_reporting():
    """Background daemon that sends trade Excel files to Telegram every 10 minutes"""
    import glob as glob_module
    
    REPORT_INTERVAL = 600
    
    def reporter_loop():
        print("[AUTO-REPORTER v11.1] Started - Sending backups every 10 minutes")
        
        while True:
            try:
                time.sleep(REPORT_INTERVAL)
                
                excel_files = glob_module.glob("trades_*.xlsx")
                
                if not excel_files:
                    print("[AUTO-REPORTER v11.1] No trade files found to backup")
                    continue
                
                current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                print(f"[AUTO-REPORTER v11.1] Sending {len(excel_files)} file(s) to Telegram...")
                
                for file_path in excel_files:
                    try:
                        filename = os.path.basename(file_path)
                        caption = f"📂 Auto-Backup: {filename} | Time: {current_time}"
                        
                        loop = asyncio.new_event_loop()
                        asyncio.set_event_loop(loop)
                        success = loop.run_until_complete(send_telegram_document(file_path, caption))
                        loop.close()
                        
                        if success:
                            print(f"[AUTO-REPORTER v11.1] ✅ Sent: {filename}")
                        else:
                            print(f"[AUTO-REPORTER v11.1] ❌ Failed: {filename}")
                            
                    except Exception as file_error:
                        print(f"[AUTO-REPORTER v11.1] Error sending {file_path}: {file_error}")
                
                print(f"[AUTO-REPORTER v11.1] Backup complete. Next backup in {REPORT_INTERVAL}s")
                
            except Exception as e:
                print(f"[AUTO-REPORTER v11.1] Reporter error: {e}")
                time.sleep(60)
    
    reporter_thread = threading.Thread(target=reporter_loop, daemon=True)
    reporter_thread.start()
    return reporter_thread


# ==========================================
# 10. BALANCE MANAGEMENT
# ==========================================
async def get_cached_balance(api, force_refresh=False):
    async with balance_lock:
        current_time = time.time()
        if not force_refresh and (current_time - state.balance_last_updated) < 60:
            return state.cached_balance, state.cached_currency

        try:
            balance_response = await api.balance()
            if 'balance' in balance_response:
                state.cached_balance = float(balance_response['balance'].get('balance', 0))
                state.cached_currency = balance_response['balance'].get('currency', 'USD')
                state.balance_last_updated = current_time
                return state.cached_balance, state.cached_currency
        except Exception as e:
            print(f"Balance fetch error: {e}")

        return state.cached_balance, state.cached_currency


# ==========================================
# 11. MARKET SCANNER v10.1 (Single Thread Discipline)
# ==========================================
async def scan_market(api, symbol):
    try:
        candles = await api.ticks_history({
            "ticks_history": symbol,
            "adjust_start_time": 1,
            "count": 100,
            "end": "latest",
            "granularity": CONFIG["TIMEFRAME_ENTRY"],
            "style": "candles"
        })

        if 'candles' not in candles:
            return None

        candle_data = candles['candles']
        if len(candle_data) < 30:
            return None

        df = pd.DataFrame(candle_data)
        df = df.rename(columns={'epoch': 'time'})
        df = calculate_indicators(df)

        if df is None:
            return None

        async with price_buffer_lock:
            if symbol in price_history_buffer:
                current_price = float(df.iloc[-1]['close'])
                price_history_buffer[symbol].append({
                    'time': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'price': current_price
                })

        direction, strength, signal_name, signal_df = determine_sniper_signal(df, symbol)

        if direction and strength >= 70:
            current_price = float(df.iloc[-1]['close'])
            adx = float(df.iloc[-1]['adx'])
            is_crash = adx >= CONFIG["ADX_CRASH_THRESHOLD"]

            strategy = "CRASH_HUNTER" if is_crash else ("TREND_SURFER" if "TREND" in signal_name else "NORMAL_PULLBACK")

            return {
                'symbol': symbol,
                'direction': direction,
                'strength': strength,
                'signal_name': signal_name,
                'current_price': current_price,
                'strategy': strategy,
                'is_crash_hunter': is_crash,
                'df': signal_df,
                'signal_time': time.time(),
                'signal_price': current_price
            }

        return None

    except Exception as e:
        if 'MarketIsClosed' not in str(e):
            print(f"Scan error for {symbol}: {e}")
        return None


async def market_scanner(api):
    """v11.1 Full Auto-Reporter Edition - ONE TRADE AT A TIME"""
    print(f"\n🔒 SINGLE THREAD SCANNER STARTED (v11.1 Full Auto-Reporter Edition)")
    print(f"Markets: {CONFIG['VOLATILITY_ASSETS']}")
    print(f"[BLACK BOX v10.0] Price buffer active - Recording last 20 ticks per symbol")

    stake = CONFIG["SNIPER_STAKE"]
    volatile_cfg = SYMBOL_CONFIG.get("R_10", DEFAULT_CONFIG)
    stable_cfg = DEFAULT_CONFIG
    print(f"[DIVIDE & RULE v10.2] Stake: ${stake:.2f}")
    print(f"   Volatile (R_10/R_25): Gap ${stake*volatile_cfg['SHADOW_GAP_PCT']:.2f} [TIGHT]")
    print(f"   Stable (R_50+): Gap ${stake*stable_cfg['SHADOW_GAP_PCT']:.2f} [WIDE]")
    print(f"[DISCIPLINE v10.1] Cooldown: {CONFIG['POST_TRADE_COOLDOWN']}s | Max Trades: 1")

    while state.daily_active:
        try:
            if len(state.active_trades) > 0:
                print(f"⏳ Trade in progress on {list(state.active_trades.keys())}... waiting")
                await asyncio.sleep(5)
                continue

            cooldown_remaining = time.time() - state.last_trade_end_time
            if state.last_trade_end_time > 0 and cooldown_remaining < CONFIG["POST_TRADE_COOLDOWN"]:
                remaining = CONFIG["POST_TRADE_COOLDOWN"] - cooldown_remaining
                print(f"☕ Cooldown... {int(remaining)}s remaining")
                await asyncio.sleep(min(10, remaining))
                continue

            if state.total_pnl >= CONFIG["DAILY_TARGET_PROFIT"]:
                print(f"🎯 DAILY TARGET REACHED: ${state.total_pnl:.2f}")
                await send_telegram(f"🎯 DAILY TARGET REACHED!\nProfit: ${state.total_pnl:.2f}")
                await asyncio.sleep(60)
                continue

            if state.total_pnl <= -CONFIG["DAILY_MAX_LOSS"]:
                print(f"🛑 DAILY MAX LOSS REACHED: ${state.total_pnl:.2f}")
                await send_telegram(f"🛑 DAILY MAX LOSS!\nLoss: ${state.total_pnl:.2f}")
                state.daily_active = False
                break

            print(f"🔍 Scanning markets... (P/L: ${state.total_pnl:.2f})")
            signals = []
            for symbol in CONFIG["VOLATILITY_ASSETS"]:
                signal = await scan_market(api, symbol)
                if signal:
                    signals.append(signal)
                    print(f"   📡 Signal found: {symbol} {signal['direction']} (Strength: {signal['strength']})")
                await asyncio.sleep(0.5)

            if not signals:
                await asyncio.sleep(10)
                continue

            signals.sort(key=lambda x: (x.get('is_crash_hunter', False), x['strength']), reverse=True)

            best_signal = signals[0]
            print(f"\n🎯 ONE BULLET: Executing BEST signal - {best_signal['symbol']} {best_signal['direction']}")

            if len(signals) > 1:
                ignored = [s['symbol'] for s in signals[1:]]
                print(f"   ⏭️ Ignored weaker signals: {ignored}")

            _, currency = await get_cached_balance(api)

            trade_manager = TradeManager(
                api=api,
                symbol=best_signal['symbol'],
                strategy_type=best_signal['strategy'],
                direction=best_signal['direction'],
                stake=CONFIG["SNIPER_STAKE"],
                duration=CONFIG["SNIPER_DURATION"],
                current_price=best_signal['current_price'],
                signal_strength=best_signal['strength'],
                signal_name=best_signal['signal_name'],
                is_crash_hunter=best_signal.get('is_crash_hunter', False),
                entry_df=best_signal.get('df'),
                signal_time=best_signal.get('signal_time'),
                signal_price=best_signal.get('signal_price')
            )

            await trade_manager.execute(currency)

            await asyncio.sleep(5)

        except Exception as e:
            print(f"Scanner error: {e}")
            traceback.print_exc()
            await asyncio.sleep(10)


# ==========================================
# 12. MAIN BOT FUNCTION
# ==========================================
async def main_bot():
    try:
        if not CONFIG["API_TOKEN"]:
            print("ERROR: DERIV_API_TOKEN not set!")
            return

        api = DerivAPI(app_id=CONFIG["APP_ID"])

        auth_response = await api.authorize({"authorize": CONFIG["API_TOKEN"]})

        if 'error' in auth_response:
            print(f"Authorization failed: {auth_response['error']}")
            return

        auth_info = auth_response.get('authorize', {})
        email = auth_info.get('email', 'Unknown')
        print(f"Authorized: {email}")

        balance, currency = await get_cached_balance(api, force_refresh=True)
        print(f"Account Balance: {currency} {balance:.2f}")

        await send_telegram(
            f"🤖 AVITS v11.1 STARTED (Full Auto-Reporter Edition)\n"
            f"Mode: SINGLE THREAD + AUTO-REPORTER\n"
            f"Balance: {currency} {balance:.2f}\n"
            f"Stake: ${CONFIG['SNIPER_STAKE']:.2f}\n"
            f"R_10/R_25: 10% Gap (Tight) | TP: 65%\n"
            f"R_50+: 35% Gap (Wide) | TP: 80%\n"
            f"Backups: Every 10 minutes\n"
            f"Markets: {', '.join(CONFIG['VOLATILITY_ASSETS'])}"
        )

        await market_scanner(api)

    except Exception as e:
        print(f"Main bot error: {e}")
        traceback.print_exc()
        raise


# ==========================================
# 13. ZOMBIE RESURRECTION LOOP
# ==========================================
def run_bot_with_resurrection():
    while True:
        try:
            print(f"\n{'='*60}")
            print("🧟 ZOMBIE MODE: Starting/Restarting bot...")
            print(f"Restart count: {state.restart_count}")
            print(f"{'='*60}\n")

            asyncio.run(main_bot())

        except KeyboardInterrupt:
            print("\n🛑 Bot stopped by user")
            break
        except Exception as e:
            state.restart_count += 1
            print(f"\n💀 BOT CRASHED: {e}")
            print(f"🔄 Restarting in 10 seconds... (Restart #{state.restart_count})")
            traceback.print_exc()
            time.sleep(10)


# ==========================================
# 14. ENTRY POINT
# ==========================================
if __name__ == "__main__":
    print_dynamic_math_check()

    init_excel()
    init_blackbox_csv()
    init_price_history_buffer()

    if run_diagnostics():
        flask_thread = threading.Thread(target=run_keep_alive, daemon=True)
        flask_thread.start()
        print("Flask keep-alive server started on port 5000")

        heartbeat_thread = threading.Thread(target=start_heartbeat, daemon=True)
        heartbeat_thread.start()
        print("Heartbeat thread started (60s interval)")

        reporter_thread = start_periodic_reporting()
        print("[AUTO-REPORTER v11.1] Started - Backups every 10 minutes to Telegram")

        run_bot_with_resurrection()
    else:
        print("Bot startup failed - check configuration")
